"""Seed Supabase with FAAC allocation and IGR data from Excel.

Usage:
  1. First run supabase/schema.sql in Supabase SQL Editor
  2. Then: python scripts/seed.py

Requires env vars:
  NEXT_PUBLIC_SUPABASE_URL=<your-url>
  NEXT_PUBLIC_SUPABASE_ANON_KEY=<your-anon-key>
"""

import os
import sys
from pathlib import Path

import openpyxl
from supabase import create_client

SUPABASE_URL = os.environ.get(
    "NEXT_PUBLIC_SUPABASE_URL",
    "https://pvguhssnvzldvnnhmoqk.supabase.co",
)
SUPABASE_KEY = os.environ.get(
    "NEXT_PUBLIC_SUPABASE_ANON_KEY",
    "sb_publishable_bbvWdXIAnnAkbuD0sEmV5A_qj1sCLFb",
)

XLSX_PATH = (
    Path(__file__).parent.parent / "FAAC_Allocation_Aggregated_2016-2026.xlsx"
)

MONTH_NAMES = [
    "january", "february", "march", "april", "may", "june",
    "july", "august", "september", "october", "november", "december",
]


def normalize_state(name: str) -> str:
    return name.strip().title()


def find_column(headers, *keywords):
    """Find column index matching all keywords."""
    for idx, h in enumerate(headers):
        if h is None:
            continue
        hl = str(h).lower().strip()
        if all(k.lower() in hl for k in keywords):
            return idx
    return None


def parse_xlsx():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    faac_rows = []
    igr_rows = []

    for sheet_name in wb.sheetnames:
        try:
            year = int(sheet_name)
        except ValueError:
            continue

        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        if not headers:
            continue

        state_col = find_column(headers, "state")
        if state_col is None:
            continue

        # Find monthly net/gross columns
        month_cols = {}
        for m in MONTH_NAMES:
            net_idx = find_column(headers, m, "net")
            gross_idx = find_column(headers, m, "gross")
            if net_idx is not None or gross_idx is not None:
                month_cols[m] = (gross_idx, net_idx)

        # Find IGR column (if exists)
        igr_col = find_column(headers, "total", "igr") or find_column(
            headers, "igr"
        )

        for row in ws.iter_rows(min_row=2, values_only=True):
            state_raw = row[state_col]
            if state_raw is None:
                continue
            state = normalize_state(str(state_raw))

            # Skip summary/total rows
            if state.lower() in ("total", "grand total", ""):
                continue

            # IGR
            if igr_col is not None and igr_col < len(row) and row[igr_col] is not None:
                try:
                    igr_val = float(str(row[igr_col]).replace(",", ""))
                    igr_rows.append({
                        "year": year,
                        "state": state,
                        "total_igr": igr_val,
                    })
                except (ValueError, TypeError):
                    pass

            # Monthly FAAC allocations
            for month_idx, month_name in enumerate(MONTH_NAMES, start=1):
                if month_name not in month_cols:
                    continue
                gross_idx, net_idx = month_cols[month_name]

                gross_val = None
                net_val = None

                if gross_idx is not None and gross_idx < len(row) and row[gross_idx] is not None:
                    try:
                        gross_val = float(str(row[gross_idx]).replace(",", ""))
                    except (ValueError, TypeError):
                        pass

                if net_idx is not None and net_idx < len(row) and row[net_idx] is not None:
                    try:
                        net_val = float(str(row[net_idx]).replace(",", ""))
                    except (ValueError, TypeError):
                        pass

                if gross_val is not None or net_val is not None:
                    faac_rows.append({
                        "year": year,
                        "state": state,
                        "month": month_idx,
                        "gross": gross_val or 0,
                        "net": net_val or 0,
                    })

    return faac_rows, igr_rows


def seed():
    print("Parsing Excel file...")
    faac_rows, igr_rows = parse_xlsx()
    print(f"Found {len(faac_rows)} FAAC rows, {len(igr_rows)} IGR rows")

    print("Connecting to Supabase...")
    supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

    # Clear existing data
    print("Clearing existing data...")
    supabase.table("faac_allocations").delete().neq("year", 0).execute()
    supabase.table("igr_data").delete().neq("year", 0).execute()

    # Insert FAAC data in batches
    print("Inserting FAAC allocations...")
    BATCH = 500
    for i in range(0, len(faac_rows), BATCH):
        batch = faac_rows[i : i + BATCH]
        res = supabase.table("faac_allocations").upsert(batch, on_conflict="year,state,month").execute()
        print(f"  FAAC batch {i // BATCH + 1}: {len(batch)} rows")

    # Insert IGR data
    print("Inserting IGR data...")
    for i in range(0, len(igr_rows), BATCH):
        batch = igr_rows[i : i + BATCH]
        res = supabase.table("igr_data").upsert(batch, on_conflict="year,state").execute()
        print(f"  IGR batch {i // BATCH + 1}: {len(batch)} rows")

    # Refresh materialized view
    print("Refreshing dashboard_summary view...")
    supabase.rpc("refresh_dashboard_summary").execute()

    print("Done! Verify by fetching from dashboard_summary.")
    count = supabase.table("dashboard_summary").select("*", count="exact").execute()
    print(f"Dashboard summary has {len(count.data)} rows")


if __name__ == "__main__":
    seed()
